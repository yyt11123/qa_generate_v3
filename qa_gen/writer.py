# -*- coding: utf-8 -*-
"""xlsx & run_log.md 写出。"""
import time
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from qa_gen.config import (
    _stats,
    INPUT_JSONL,
    OUT_XLSX,
    OUT_LOG,
    LLM_MODEL,
    EMB_MODEL,
    MAX_LLM_CALLS,
    MAX_EMB_BATCHES,
)


# -------- 写出 xlsx --------
def write_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头：第1行
    ws.cell(1, 1, None)
    ws.cell(1, 2, "question（咨询问题）")
    ws.cell(1, 3, "answers（预期回答）")
    ws.cell(1, 4, "supporting facts（支撑信息）")
    ws.cell(1, 5, None)
    ws.cell(1, 6, None)
    ws.cell(1, 7, "gold_chunk_ids")
    ws.cell(1, 8, "answer_type")
    ws.cell(1, 9, "intent")
    # 第2行
    ws.cell(2, 1, "分类")
    ws.cell(2, 2, None)
    ws.cell(2, 3, None)
    ws.cell(2, 4, "document（文件名称）")
    ws.cell(2, 5, "page（所在页码）")
    ws.cell(2, 6, "text/img/table（支撑文本/图片/表格）")
    ws.cell(2, 7, None)
    ws.cell(2, 8, None)
    ws.cell(2, 9, None)

    # 模板里的合并：B1:B2, C1:C2, D1:F1（supporting facts 跨D-F），A1:A2 不合并（A1空、A2写"分类"）
    # 参照模板：D1 = supporting facts 跨 D-F，第二行 D2/E2/F2 是子标题
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)  # B1:B2
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)  # C1:C2
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)  # D1:F1
    # 扩展列：G/H/I 不合并

    # 数据：按"分类"分组，连续同分类的行只在首行写分类，下面留空
    # 先按 category 排序：产品 → 健康核保 → 拒答（unanswerable 算入产品）
    cat_order = {"产品": 0, "健康核保": 1}
    rows_sorted = sorted(
        rows,
        key=lambda r: (
            cat_order.get(r["category"], 9),
            1 if r["answer_type"] == "unanswerable" else 0,
            r["intent"],
            r.get("task_key", ""),
        ),
    )

    start_row = 3
    last_cat = None
    cat_start_row = None
    for i, r in enumerate(rows_sorted):
        row_idx = start_row + i
        if r["category"] != last_cat:
            # 上一组合并 A 列（如果跨 ≥ 2 行）
            if last_cat is not None and cat_start_row is not None and (row_idx - 1) > cat_start_row:
                ws.merge_cells(start_row=cat_start_row, start_column=1, end_row=row_idx - 1, end_column=1)
            ws.cell(row_idx, 1, r["category"])
            cat_start_row = row_idx
            last_cat = r["category"]
        ws.cell(row_idx, 2, r["question"])
        ws.cell(row_idx, 3, r["answer"])
        ws.cell(row_idx, 4, r["document"])
        ws.cell(row_idx, 5, r["page"])
        ws.cell(row_idx, 6, r["supporting_text"])
        ws.cell(row_idx, 7, r["gold_chunk_ids"])
        ws.cell(row_idx, 8, r["answer_type"])
        ws.cell(row_idx, 9, r["intent"])
    # 末尾分组合并
    if cat_start_row is not None and (start_row + len(rows_sorted) - 1) > cat_start_row:
        ws.merge_cells(start_row=cat_start_row, start_column=1,
                       end_row=start_row + len(rows_sorted) - 1, end_column=1)

    # 列宽
    widths = [10, 30, 50, 30, 8, 50, 30, 14, 10]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in range(1, 10):
        for row in range(1, start_row + len(rows_sorted)):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 10):
        for row in (1, 2):
            ws.cell(row, col).font = Font(bold=True)
    wb.save(path)


# -------- 写出运行日志 --------
def write_log(rows, path, drop_stats):
    cnt_intent = Counter(r["intent"] for r in rows)
    cnt_atype = Counter(r["answer_type"] for r in rows)
    cnt_cat = Counter(r["category"] for r in rows)
    cnt_section = Counter(r["section_title"] for r in rows if r["answer_type"] != "unanswerable")
    n_unanswer = sum(1 for r in rows if r["answer_type"] == "unanswerable")

    lines = []
    lines.append("# 愛唯守 QA 测评集生成 · 运行日志\n")
    lines.append(f"- 生成时间：{time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- 数据源：`{INPUT_JSONL}`（250 chunk）")
    lines.append(f"- 生成 LLM：`{LLM_MODEL}`；Embedding：`{EMB_MODEL}`")
    lines.append(f"- 输出：`{OUT_XLSX}`、`{OUT_LOG}`\n")

    lines.append("## 总量与分布")
    lines.append(f"- 最终条数：**{len(rows)}**（其中负样本 unanswerable {n_unanswer} 条）")
    lines.append("- 按分类（A 列）：")
    for k, v in cnt_cat.most_common():
        lines.append(f"  - {k}：{v}")
    lines.append("- 按 intent（I 列）：")
    for k, v in cnt_intent.most_common():
        lines.append(f"  - {k}：{v}")
    lines.append("- 按 answer_type（H 列）：")
    for k, v in cnt_atype.most_common():
        lines.append(f"  - {k}：{v}")
    lines.append(f"- 同 section_title 出题分布（仅可答题，最多 {max(cnt_section.values()) if cnt_section else 0} 条/桶）：")
    for k, v in cnt_section.most_common():
        lines.append(f"  - {v}：{k}")

    lines.append("\n## 调用统计")
    lines.append(f"- LLM (`{LLM_MODEL}`) 调用次数：{_stats['llm_calls']}（上限 {MAX_LLM_CALLS}），错误 {_stats['llm_errors']}")
    lines.append(f"- Embedding (`{EMB_MODEL}`) 批次：{_stats['emb_batches']}（上限 {MAX_EMB_BATCHES}），错误 {_stats['emb_errors']}")
    lines.append(f"- 重试次数：{_stats['retries']}；丢弃条数：{_stats['dropped']}；去重删除：{drop_stats.get('dedup_dropped', 0)}")

    lines.append("\n## 按规格书做的默认假设（§12.1）")
    if _stats["log_assumptions"]:
        for a in _stats["log_assumptions"]:
            lines.append(f"- {a}")
    else:
        lines.append("- （无）")

    lines.append("\n## 关键事件日志")
    for ev in _stats["log_events"]:
        lines.append(f"- {ev}")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
