# -*- coding: utf-8 -*-
"""
按 QA生成规格书_愛唯守.md 的 §8 / §12 执行：
LLM (qwen-plus) 生成 + Embedding (text-embedding-v4) 语义去重，
从单一 jsonl 生成约 40 条 QA，输出到 ./output/。
"""
import os, sys, json, re, time, math, random
from collections import defaultdict, Counter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import numpy as np

# -------- 配置 --------
INPUT_JSONL = "愛唯守危疾保障_产品彩页_paged(1).jsonl"
OUT_DIR     = "./output"
OUT_XLSX    = os.path.join(OUT_DIR, "愛唯守_QA测评集.xlsx")
OUT_LOG     = os.path.join(OUT_DIR, "run_log.md")
DOC_NAME    = "愛唯守危疾保障.pdf"

LLM_MODEL = "qwen-plus"
EMB_MODEL = "text-embedding-v4"

# 上限（§12 防空转）
MAX_LLM_CALLS  = 200
MAX_EMB_BATCHES = 10
MAX_RETRIES_PER_ITEM = 2
DEDUP_SIM_THRESHOLD = 0.85

client = OpenAI(
    api_key=os.getenv("DASHSCOPE_API_KEY"),
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)

# 调用计数（全局）
_stats = {
    "llm_calls": 0, "llm_errors": 0,
    "emb_batches": 0, "emb_errors": 0,
    "retries": 0, "dropped": 0,
    "log_assumptions": [],
    "log_events": [],
}

def log_event(msg):
    print(msg)
    _stats["log_events"].append(msg)


# -------- 工具：API 调用，含指数退避 --------
def gen(messages, retries=5):
    if _stats["llm_calls"] >= MAX_LLM_CALLS:
        log_event(f"[STOP] LLM 调用上限 {MAX_LLM_CALLS} 已达，跳过此次生成。")
        return None
    delay = 1.0
    for attempt in range(retries):
        try:
            _stats["llm_calls"] += 1
            r = client.chat.completions.create(
                model=LLM_MODEL,
                messages=messages,
                response_format={"type": "json_object"},
                extra_body={"enable_thinking": False},
                temperature=0.4,
                timeout=60,
            )
            txt = r.choices[0].message.content
            try:
                return json.loads(txt)
            except json.JSONDecodeError:
                m = re.search(r"\{.*\}", txt, re.S)
                return json.loads(m.group(0)) if m else None
        except Exception as e:
            _stats["llm_errors"] += 1
            msg = str(e)
            unrecoverable = any(k in msg for k in ["401", "Authentication", "InvalidApiKey", "model not found", "ModelNotFound", "402", "PermissionDenied", "AccountFlowControl"])
            if unrecoverable:
                log_event(f"[FATAL] LLM 不可自愈错误：{msg}")
                raise
            log_event(f"[WARN] LLM 调用失败({attempt+1}/{retries}): {msg[:160]}; 退避 {delay:.1f}s")
            time.sleep(delay)
            delay *= 2
    log_event("[ERROR] LLM 多次重试仍失败，返回 None。")
    return None


def embed(texts, retries=5):
    """text-embedding-v4 每次最多 10 条，按 ≤10 切分批次循环调用，结果拼接。
    每个子批计入 emb_batches 并遵守 MAX_EMB_BATCHES。"""
    if not texts:
        return []
    BATCH_SIZE = 10
    all_vecs = []
    for s in range(0, len(texts), BATCH_SIZE):
        sub = texts[s:s + BATCH_SIZE]
        if _stats["emb_batches"] >= MAX_EMB_BATCHES:
            log_event(f"[STOP] embedding 批次上限 {MAX_EMB_BATCHES} 已达；剩余 {len(texts)-s} 条未向量化。")
            return None
        delay = 1.0
        sub_vecs = None
        for attempt in range(retries):
            try:
                _stats["emb_batches"] += 1
                r = client.embeddings.create(
                    model=EMB_MODEL, input=sub,
                    dimensions=1024, encoding_format="float", timeout=60,
                )
                sub_vecs = [d.embedding for d in r.data]
                break
            except Exception as e:
                _stats["emb_errors"] += 1
                msg = str(e)
                unrecoverable = any(k in msg for k in ["401", "Authentication", "InvalidApiKey", "model not found", "ModelNotFound", "402", "PermissionDenied"])
                if unrecoverable:
                    log_event(f"[FATAL] EMB 不可自愈错误：{msg}")
                    raise
                log_event(f"[WARN] EMB 调用失败({attempt+1}/{retries}): {msg[:160]}; 退避 {delay:.1f}s")
                time.sleep(delay)
                delay *= 2
        if sub_vecs is None:
            return None
        all_vecs.extend(sub_vecs)
    return all_vecs


# -------- 数据加载 --------
def load_chunks(path):
    chunks = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                chunks.append(json.loads(line))
    by_id = {c["chunk_id"]: c for c in chunks}
    return chunks, by_id


# -------- F 列子串校验（去空白后比对） --------
def _norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s))


def clean_question(q):
    """问题清洗：删除中英文括号及括号内内容；保证只在末尾保留一个问号。"""
    if not q:
        return q
    q = re.sub(r"[（(][^）)]*[）)]", "", q)
    q = q.replace("?", "？")
    if q.count("？") >= 2:
        parts = q.split("？")
        head = "、".join(p for p in parts[:-1] if p != "")
        tail = parts[-1]
        q = head + "？" + (tail if tail else "")
    q = re.sub(r"、+", "、", q)
    q = re.sub(r"\s{2,}", " ", q).strip()
    return q


def find_supporting_chunk(supporting_text, candidate_chunks):
    """返回 supporting_text 是其 content 子串的第一个 chunk_id；找不到返回 None。"""
    target = _norm(supporting_text)
    if not target:
        return None
    for c in candidate_chunks:
        if target in _norm(c.get("content") or ""):
            return c["chunk_id"]
    return None


# -------- 关键短语提取与 G 列回填 --------
def extract_keyphrases(text):
    """从 answer/supporting_text 抽取关键短语：数字+量词、百分比、显式数字。"""
    if not text:
        return []
    phrases = set()
    for m in re.findall(r"\d{1,4}(?:,\d{3})*(?:\.\d+)?\s*(?:%|％|港元|美元|澳門幣|個月|個|个月|个|歲|岁|次|年|天|日|份|港币|美金|港|公里|周|週)", text):
        phrases.add(m.strip())
    for m in re.findall(r"\d{1,4}(?:,\d{3})*(?:\.\d+)?\s*[%％]", text):
        phrases.add(m.strip())
    # 中文数字+量词的稳健形式
    for m in re.findall(r"\d{1,4}\s*[次年月日歲岁个個%％]", text):
        phrases.add(m.strip())
    # 关键中文短语
    for kw in ["最多索償", "最多索偿", "最多", "等候期", "保額", "保额", "保單貨幣", "投保", "繳付", "缴付", "額外保障", "終期紅利", "现金价值"]:
        if kw in text:
            phrases.add(kw)
    return [p for p in phrases if len(p) >= 1]


def fill_gold_ids(supporting_text, primary_id, all_chunks):
    """G 列：保留所有 content（空白归一化后）包含 supporting_text 的 chunk。
    必须含 primary_id；排除 doc_summary（除非它就是 primary）；按 chunk_index 排序。"""
    nf = _norm(supporting_text)
    hits = []
    if nf:
        for c in all_chunks:
            if (c["chunk_type"] == "doc_summary") and (c["chunk_id"] != primary_id):
                continue
            if nf in _norm(c.get("content") or ""):
                hits.append(c)
    ids = {c["chunk_id"] for c in hits}
    ids.add(primary_id)
    idx_of = {c["chunk_id"]: c["chunk_index"] for c in all_chunks}
    return sorted(ids, key=lambda cid: idx_of.get(cid, 9999))


# -------- 余弦相似度 --------
def cos_sim_matrix(vecs):
    m = np.array(vecs, dtype=np.float32)
    norm = np.linalg.norm(m, axis=1, keepdims=True) + 1e-12
    m = m / norm
    return m @ m.T


# -------- 任务桶定义（§4 配额，覆盖 §5 的多个 section_title） --------
# 每个任务：intent / sub_intent / category / answer_type_hint / chunk_ids / hint
BUCKETS = [
    # === A 查详情：保障范围/特性 (产品, 14) ===
    {"key":"A1", "intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0003","9260cfa4_chunk_0013"],
     "hint":"3大早期风险状况（心肌病早期擴張型 / 大腸息肉腺瘤性 / 原位癌）的范围"},
    {"key":"A2", "intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0013"],
     "hint":"早期风险守护保险赔偿支付次数（每种早期风险只赔1次，合共2次）"},
    {"key":"A3", "intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0008"],
     "hint":"持续癌症保险赔偿是什么 / 怎么赔（每月保额5%、长达100个月）"},
    {"key":"A4", "intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0009","9260cfa4_chunk_0012"],
     "hint":"严重认知障碍症照顾者年金（每年保额6%，至100岁）"},
    {"key":"A5", "intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0014"],
     "hint":"严重疾病多重保险赔偿的覆盖（直至85岁，每次100%保额）"},
    {"key":"A6", "intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0015","9260cfa4_chunk_0017"],
     "hint":"持续癌症保险赔偿的等候期（短至1年）/ 选项条件"},
    {"key":"A7", "intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0025"],
     "hint":"教育特殊支援（6-18岁、保额5%、只索偿1次）"},
    {"key":"A8", "intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0032"],
     "hint":"愛滿途支援計劃 / 一系列增值服务（预防 / 治疗 / 康复）"},
    {"key":"A9", "intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0031"],
     "hint":"增值權益附加條款（额外保费 + 保额每年自动增加）"},
    {"key":"A10","intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0031","9260cfa4_chunk_0206"],
     "hint":"延長寬限期保障（特定事件：结婚/生育/失业/离婚 → 最多365日）"},
    {"key":"A11","intent":"查详情", "category":"产品", "answer_type":"multi_chunk",
     "chunk_ids":["9260cfa4_chunk_0028","9260cfa4_chunk_0192"],
     "hint":"额外保障规则（首10个保单年度内额外50%保额）"},
    {"key":"A12","intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0027"],
     "hint":"未成年保单的额外保障（终身）"},
    {"key":"A13","intent":"查详情", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0026"],
     "hint":"一站式财富累积（保证现金价值、终期红利）"},
    {"key":"A14","intent":"查详情", "category":"产品", "answer_type":"multi_chunk",
     "chunk_ids":["9260cfa4_chunk_0203","9260cfa4_chunk_0205","9260cfa4_chunk_0206"],
     "hint":"保单持有人身故保费豁免的核心条件（持有人保单日期年龄50岁或以下、生效满2年、75岁或之前身故）"},

    # === B 要数字 (产品, 8) ===
    {"key":"B1","intent":"要数字", "category":"产品", "answer_type":"table_lookup",
     "chunk_ids":["9260cfa4_chunk_0007"],
     "hint":"严重疾病多重保险赔偿合共最多多少次（连同严重疾病保险赔偿合共高达9次）"},
    {"key":"B2","intent":"要数字", "category":"产品", "answer_type":"table_lookup",
     "chunk_ids":["9260cfa4_chunk_0007"],
     "hint":"癌症最多索偿几次（最多5次）"},
    {"key":"B3","intent":"要数字", "category":"产品", "answer_type":"table_lookup",
     "chunk_ids":["9260cfa4_chunk_0007"],
     "hint":"心脏病发作及中风最多索偿几次（最多2次）"},
    {"key":"B4","intent":"要数字", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0008","9260cfa4_chunk_0174"],
     "hint":"持续癌症保险赔偿每月赔多少（每月保额的5%）"},
    {"key":"B5","intent":"要数字", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0008","9260cfa4_chunk_0174"],
     "hint":"持续癌症保险赔偿最长多少个月（长达100个月）"},
    {"key":"B6","intent":"要数字", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0009","9260cfa4_chunk_0012"],
     "hint":"认知障碍症照顾者年金每年多少（保额的6%）"},
    {"key":"B7","intent":"要数字", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0009","9260cfa4_chunk_0012"],
     "hint":"认知障碍症照顾者年金支付到多少岁（100岁）"},
    {"key":"B8","intent":"要数字", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0014"],
     "hint":"总保障最高多少（保额1000%）"},

    # === C 资格门槛 (产品, 4) ===
    {"key":"C1","intent":"资格门槛", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0170"],
     "hint":"非严重疾病保险赔偿的货币上限（港元/澳门币/美元三种货币的具体数字）"},
    {"key":"C2","intent":"资格门槛", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0215"],
     "hint":"澳门保单的货币选择（澳门币或其他可供选择的货币）"},
    {"key":"C3","intent":"资格门槛", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0223"],
     "hint":"暂停缴付保费的后果（宽限期31日 / 保单可能终止）"},
    {"key":"C4","intent":"资格门槛", "category":"产品", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0203"],
     "hint":"持有人身故保费豁免的年龄要求（保单日期当日50岁或以下）"},

    # === D 查覆盖：受保疾病 (健康核保, 5) ===
    {"key":"D1","intent":"查覆盖", "category":"健康核保", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0013","9260cfa4_chunk_0004"],
     "hint":"产品总共保障多少种疾病（135种非严重至严重疾病）"},
    {"key":"D2","intent":"查覆盖", "category":"健康核保", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0004"],
     "hint":"严重疾病多少种（63种）"},
    {"key":"D3","intent":"查覆盖", "category":"健康核保", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0004"],
     "hint":"非严重疾病多少种（72种）"},
    {"key":"D4","intent":"查覆盖", "category":"健康核保", "answer_type":"single_fact",
     "chunk_ids":["9260cfa4_chunk_0004","9260cfa4_chunk_0013"],
     "hint":"重大手术覆盖多少种（51种）"},
    {"key":"D5","intent":"查覆盖", "category":"健康核保", "answer_type":"table_lookup",
     "chunk_ids":["9260cfa4_chunk_0011","9260cfa4_chunk_0087"],
     "hint":"儿童非严重疾病多少种（15种）"},

    # === E 分人群保障 (产品, 3) ===
    {"key":"E1","intent":"查详情", "category":"产品", "answer_type":"multi_chunk",
     "chunk_ids":["9260cfa4_chunk_0010","9260cfa4_chunk_0021"],
     "hint":"愛寶保对孕妇 / 宝宝的保障（孕期18周起、妊娠并发症保障、产后抑郁）"},
    {"key":"E2","intent":"查详情", "category":"产品", "answer_type":"multi_chunk",
     "chunk_ids":["9260cfa4_chunk_0012"],
     "hint":"年长人士的保障（中度严重至严重认知障碍症 / 早期检测）"},
    {"key":"E3","intent":"查详情", "category":"产品", "answer_type":"multi_chunk",
     "chunk_ids":["9260cfa4_chunk_0011","9260cfa4_chunk_0025"],
     "hint":"儿童及照顾者的保障（15种儿童非严重疾病 / 教育特殊支援）"},
]

# 负样本（拒答）
NEG_BUCKETS = [
    {"key":"F1", "topic":"预缴保费优惠 / 折扣力度",
     "question":"爱唯守现在有预缴保费优惠吗？力度多少？"},
    {"key":"F2", "topic":"与保诚等他司危疾产品对比",
     "question":"爱唯守和保诚的危疾产品哪个保得多？"},
    {"key":"F3", "topic":"具体年龄 / 保额下的保费报价",
     "question":"35岁女性买50万保额，爱唯守一年保费多少？"},
    {"key":"F4", "topic":"退保现金价值具体表",
     "question":"爱唯守第10年退保现金价值是多少？"},
    {"key":"F5", "topic":"理赔时效 / 多少个工作日到账",
     "question":"爱唯守理赔多久到账，需要几个工作日？"},
    {"key":"F6", "topic":"在哪里购买 / 投保门店地址",
     "question":"爱唯守在哪里可以投保，有线下门店吗？"},
]


# -------- LLM Prompt 模板 --------
SYS_PROMPT = (
    "你是保险测评 QA 出题助手。基于「我提供的若干 chunk 原文」生成 1 条 QA。\n"
    "硬性约束（违反即视为失败）：\n"
    "1) 只能根据我给出的原文作答，禁止使用模型自身知识或跨文档补充；\n"
    "2) 问题写【简体中文】、短问、关键词式，少于 30 字（最多 35 字），一题一问；\n"
    "3) 问题以「爱唯守」或「安盛爱唯守」起手；问题不要带客户场景（除非任务明确要求）；\n"
    "4) **question 中严禁出现括号（）()或任何答案/提示内容**：题干只能是问题本身，不能携带括号备注、不能把答案的数字/百分比/术语直接复述进题干；\n"
    "5) answer 与 supporting_text 的字形必须与所选 chunk 的字形一致："
    "如果 chunk 内容是繁体则保持繁体，简体则简体；【不要做简繁互转】；\n"
    "6) supporting_text 必须是其中【某一个】chunk content 的【逐字子串】（不要改字、不要拼接）；\n"
    "7) supporting_text 长度 30~120 字之间，刚好覆盖回答依据；\n"
    "8) answer 长度 15~80 字之间，可对原文做最小整理（保留数字/百分号/单位）；\n"
    "输出严格 JSON：{\"question\":\"...\",\"answer\":\"...\",\"supporting_text\":\"...\",\"primary_chunk_id\":\"...\"}。\n"
    "primary_chunk_id 取你 supporting_text 的来源 chunk 的 chunk_id（必须来自我列出的 chunk_ids）。"
)


def build_user_prompt(task, chunks_for_task):
    parts = [
        f"任务：{task['hint']}",
        f"答案类型倾向：{task['answer_type']}（single_fact / multi_chunk / table_lookup）",
        f"问题分类：{task['category']}（A 列固定为该值）",
        f"意图：{task['intent']}（I 列固定为该值）",
        "",
        "可用 chunk（content 已与原 jsonl 一致，注意字形）：",
    ]
    for c in chunks_for_task:
        page = c.get("page_start") or "（无）"
        parts.append(f"### chunk_id = {c['chunk_id']} | section = {c['section_title']} | page = {page} | type = {c['chunk_type']}")
        parts.append(c.get("content") or "")
        parts.append("")
    parts.append("请针对【任务】写 1 条 QA，严格按系统要求输出 JSON。")
    return "\n".join(parts)


def gen_one_qa(task, by_id, all_chunks):
    chunks_for_task = [by_id[cid] for cid in task["chunk_ids"] if cid in by_id]
    if not chunks_for_task:
        log_event(f"[SKIP] 任务 {task['key']} 找不到 chunk_ids，跳过。")
        return None
    user_prompt = build_user_prompt(task, chunks_for_task)
    last_err = None
    for attempt in range(MAX_RETRIES_PER_ITEM + 1):
        out = gen([
            {"role": "system", "content": SYS_PROMPT},
            {"role": "user", "content": user_prompt},
        ])
        if not isinstance(out, dict):
            last_err = "LLM 返回非 dict"
            _stats["retries"] += int(attempt > 0)
            continue
        q = (out.get("question") or "").strip()
        a = (out.get("answer") or "").strip()
        st = (out.get("supporting_text") or "").strip()
        pcid = (out.get("primary_chunk_id") or "").strip()
        # 写出前先清洗 question：去括号 + 多问号合并
        q = clean_question(q)
        if not (q and a and st):
            last_err = "字段缺失"
            _stats["retries"] += int(attempt > 0)
            continue
        # 校验：question < 50 字、简体；这里不强转，直接长度检查
        if len(q) > 50:
            last_err = f"question 太长 ({len(q)})"
            _stats["retries"] += int(attempt > 0)
            continue
        # 校验：supporting_text 是否为某 candidate chunk 的子串
        primary_id = find_supporting_chunk(st, chunks_for_task)
        if not primary_id:
            last_err = f"supporting_text 不是任一 chunk 的子串"
            _stats["retries"] += int(attempt > 0)
            continue
        # 用 LLM 报告的 primary_chunk_id 兜底（如有效则优先）
        if pcid in [c["chunk_id"] for c in chunks_for_task]:
            # 但仍要求 supporting_text 是该 chunk 的子串
            if _norm(st) in _norm(by_id[pcid].get("content") or ""):
                primary_id = pcid
        page = by_id[primary_id].get("page_start")
        # 回填 G 列：F 子串匹配（不再用关键词灌水）
        gold_ids = fill_gold_ids(st, primary_id, all_chunks)
        # 收紧后若只剩 1 个 chunk，但任务声明 multi_chunk，则回退为 single_fact
        effective_atype = task["answer_type"]
        if len(gold_ids) <= 1 and effective_atype == "multi_chunk":
            effective_atype = "single_fact"
        return {
            "category": task["category"],
            "question": q,
            "answer": a,
            "document": DOC_NAME,
            "page": page if page is not None else "",
            "supporting_text": st,
            "gold_chunk_ids": ";".join(gold_ids),
            "answer_type": effective_atype,
            "intent": task["intent"],
            "primary_chunk_id": primary_id,
            "section_title": by_id[primary_id]["section_title"],
            "task_key": task["key"],
        }
    log_event(f"[DROP] 任务 {task['key']} 重试 {MAX_RETRIES_PER_ITEM} 次仍失败：{last_err}")
    _stats["dropped"] += 1
    return None


def gen_negatives():
    """直接构造负样本：按规格 §10 示例，不需要调 LLM。"""
    rows = []
    for nb in NEG_BUCKETS:
        rows.append({
            "category": "产品",
            "question": clean_question(nb["question"]),
            "answer": "該產品彩頁未提供該信息，無法從本文檔回答。",
            "document": DOC_NAME,
            "page": "",
            "supporting_text": "",
            "gold_chunk_ids": "",
            "answer_type": "unanswerable",
            "intent": "拒答",
            "primary_chunk_id": "",
            "section_title": "[NEGATIVE]",
            "task_key": nb["key"],
        })
    return rows


# -------- 语义去重（embedding） --------
def dedup_by_embedding(rows):
    """同 section_title 桶内 cos>=0.85 视为重复，保留信息更全（answer 更长）的一条。"""
    answerable = [r for r in rows if r["answer_type"] != "unanswerable"]
    if not answerable:
        return rows
    # 把所有 question 一起向量化（≤1 批，因为 ≤40 条）
    texts = [r["question"] for r in answerable]
    vecs = embed(texts)
    if vecs is None:
        log_event("[WARN] embedding 失败，跳过语义去重。")
        return rows
    sim = cos_sim_matrix(vecs)
    # 按 section_title 桶内做两两比较
    bucket = defaultdict(list)
    for i, r in enumerate(answerable):
        bucket[r["section_title"]].append(i)
    drop = set()
    for sec, idxs in bucket.items():
        for ii in range(len(idxs)):
            for jj in range(ii + 1, len(idxs)):
                a, b = idxs[ii], idxs[jj]
                if a in drop or b in drop:
                    continue
                if sim[a, b] >= DEDUP_SIM_THRESHOLD:
                    # 保留 answer 更长的（信息更全）；若一样长保留较短 question
                    if len(answerable[a]["answer"]) >= len(answerable[b]["answer"]):
                        drop.add(b)
                    else:
                        drop.add(a)
    if drop:
        log_event(f"[DEDUP] 同 section 内 cos≥{DEDUP_SIM_THRESHOLD} 删除 {len(drop)} 条。")
    kept_answerable = [r for i, r in enumerate(answerable) if i not in drop]
    negatives = [r for r in rows if r["answer_type"] == "unanswerable"]
    return kept_answerable + negatives


# -------- 校验 §9 --------
def validate_rows(rows, by_id, all_chunks):
    issues = []
    short_count = 0
    for idx, r in enumerate(rows):
        if r["answer_type"] == "unanswerable":
            if r["gold_chunk_ids"]:
                issues.append((idx, "unanswerable 行 G 列应为空"))
            continue
        # G 列 chunk 都存在
        gids = [g for g in r["gold_chunk_ids"].split(";") if g]
        for g in gids:
            if g not in by_id:
                issues.append((idx, f"gold chunk_id 不存在：{g}"))
        # F 列必须是某 gold chunk 的子串
        f_ok = False
        for g in gids:
            if _norm(r["supporting_text"]) in _norm(by_id[g].get("content") or ""):
                f_ok = True
                break
        if not f_ok:
            issues.append((idx, "F 列不是任一 gold chunk 的子串"))
        # document 名
        if r["document"] != DOC_NAME:
            issues.append((idx, f"document 列错：{r['document']}"))
        # question 长度 / 简体 / 一题一问（粗略）
        if len(r["question"]) > 50:
            issues.append((idx, f"question 超过 50 字：{len(r['question'])}"))
        if len(r["question"]) < 30:
            short_count += 1
        # 一题一问：清洗后应只剩末尾 1 个问号；含括号视为不合格
        cleaned = clean_question(r["question"])
        if cleaned != r["question"]:
            issues.append((idx, "question 含括号或多问号（应已被 clean_question 处理）"))
        if (cleaned.count("？") + cleaned.count("?")) > 1:
            issues.append((idx, "question 含多个问号（清洗后仍 >1）"))
    # ≥90% 短问
    answerable_n = sum(1 for r in rows if r["answer_type"] != "unanswerable")
    short_ratio = short_count / max(1, answerable_n)
    return issues, short_ratio


# -------- 写出 xlsx --------
def write_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头：第1行
    ws.cell(1, 1, None)
    ws.cell(1, 2, "question（咨询问题）")
    ws.cell(1, 3, "answers（预期回答）")
    ws.cell(1, 4, "supporting facts（支撑信息）")
    ws.cell(1, 5, None)
    ws.cell(1, 6, None)
    ws.cell(1, 7, "gold_chunk_ids")
    ws.cell(1, 8, "answer_type")
    ws.cell(1, 9, "intent")
    # 第2行
    ws.cell(2, 1, "分类")
    ws.cell(2, 2, None)
    ws.cell(2, 3, None)
    ws.cell(2, 4, "document（文件名称）")
    ws.cell(2, 5, "page（所在页码）")
    ws.cell(2, 6, "text/img/table（支撑文本/图片/表格）")
    ws.cell(2, 7, None)
    ws.cell(2, 8, None)
    ws.cell(2, 9, None)

    # 模板里的合并：B1:B2, C1:C2, D1:F1（supporting facts 跨D-F），A1:A2 不合并（A1空、A2写"分类"）
    # 参照模板：D1 = supporting facts 跨 D-F，第二行 D2/E2/F2 是子标题
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)  # B1:B2
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)  # C1:C2
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)  # D1:F1
    # 扩展列：G/H/I 不合并

    # 数据：按"分类"分组，连续同分类的行只在首行写分类，下面留空
    # 先按 category 排序：产品 → 健康核保 → 拒答（unanswerable 算入产品）
    cat_order = {"产品": 0, "健康核保": 1}
    rows_sorted = sorted(
        rows,
        key=lambda r: (
            cat_order.get(r["category"], 9),
            1 if r["answer_type"] == "unanswerable" else 0,
            r["intent"],
            r.get("task_key", ""),
        ),
    )

    start_row = 3
    last_cat = None
    cat_start_row = None
    for i, r in enumerate(rows_sorted):
        row_idx = start_row + i
        if r["category"] != last_cat:
            # 上一组合并 A 列（如果跨 ≥ 2 行）
            if last_cat is not None and cat_start_row is not None and (row_idx - 1) > cat_start_row:
                ws.merge_cells(start_row=cat_start_row, start_column=1, end_row=row_idx - 1, end_column=1)
            ws.cell(row_idx, 1, r["category"])
            cat_start_row = row_idx
            last_cat = r["category"]
        ws.cell(row_idx, 2, r["question"])
        ws.cell(row_idx, 3, r["answer"])
        ws.cell(row_idx, 4, r["document"])
        ws.cell(row_idx, 5, r["page"])
        ws.cell(row_idx, 6, r["supporting_text"])
        ws.cell(row_idx, 7, r["gold_chunk_ids"])
        ws.cell(row_idx, 8, r["answer_type"])
        ws.cell(row_idx, 9, r["intent"])
    # 末尾分组合并
    if cat_start_row is not None and (start_row + len(rows_sorted) - 1) > cat_start_row:
        ws.merge_cells(start_row=cat_start_row, start_column=1,
                       end_row=start_row + len(rows_sorted) - 1, end_column=1)

    # 列宽
    widths = [10, 30, 50, 30, 8, 50, 30, 14, 10]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in range(1, 10):
        for row in range(1, start_row + len(rows_sorted)):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 10):
        for row in (1, 2):
            ws.cell(row, col).font = Font(bold=True)
    wb.save(path)


# -------- 写出运行日志 --------
def write_log(rows, path, drop_stats):
    cnt_intent = Counter(r["intent"] for r in rows)
    cnt_atype = Counter(r["answer_type"] for r in rows)
    cnt_cat = Counter(r["category"] for r in rows)
    cnt_section = Counter(r["section_title"] for r in rows if r["answer_type"] != "unanswerable")
    n_unanswer = sum(1 for r in rows if r["answer_type"] == "unanswerable")

    lines = []
    lines.append("# 愛唯守 QA 测评集生成 · 运行日志\n")
    lines.append(f"- 生成时间：{time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- 数据源：`{INPUT_JSONL}`（250 chunk）")
    lines.append(f"- 生成 LLM：`{LLM_MODEL}`；Embedding：`{EMB_MODEL}`")
    lines.append(f"- 输出：`{OUT_XLSX}`、`{OUT_LOG}`\n")

    lines.append("## 总量与分布")
    lines.append(f"- 最终条数：**{len(rows)}**（其中负样本 unanswerable {n_unanswer} 条）")
    lines.append("- 按分类（A 列）：")
    for k, v in cnt_cat.most_common():
        lines.append(f"  - {k}：{v}")
    lines.append("- 按 intent（I 列）：")
    for k, v in cnt_intent.most_common():
        lines.append(f"  - {k}：{v}")
    lines.append("- 按 answer_type（H 列）：")
    for k, v in cnt_atype.most_common():
        lines.append(f"  - {k}：{v}")
    lines.append(f"- 同 section_title 出题分布（仅可答题，最多 {max(cnt_section.values()) if cnt_section else 0} 条/桶）：")
    for k, v in cnt_section.most_common():
        lines.append(f"  - {v}：{k}")

    lines.append("\n## 调用统计")
    lines.append(f"- LLM (`{LLM_MODEL}`) 调用次数：{_stats['llm_calls']}（上限 {MAX_LLM_CALLS}），错误 {_stats['llm_errors']}")
    lines.append(f"- Embedding (`{EMB_MODEL}`) 批次：{_stats['emb_batches']}（上限 {MAX_EMB_BATCHES}），错误 {_stats['emb_errors']}")
    lines.append(f"- 重试次数：{_stats['retries']}；丢弃条数：{_stats['dropped']}；去重删除：{drop_stats.get('dedup_dropped', 0)}")

    lines.append("\n## 按规格书做的默认假设（§12.1）")
    if _stats["log_assumptions"]:
        for a in _stats["log_assumptions"]:
            lines.append(f"- {a}")
    else:
        lines.append("- （无）")

    lines.append("\n## 关键事件日志")
    for ev in _stats["log_events"]:
        lines.append(f"- {ev}")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


# -------- 主流程 --------
def main():
    if not os.getenv("DASHSCOPE_API_KEY"):
        log_event("[FATAL] 环境变量 DASHSCOPE_API_KEY 未设置。")
        sys.exit(1)
    os.makedirs(OUT_DIR, exist_ok=True)

    log_event("[STEP1] 加载 jsonl 与建索引")
    chunks, by_id = load_chunks(INPUT_JSONL)
    log_event(f"  共 {len(chunks)} 个 chunk")

    # 默认假设记录（§12 要求）
    _stats["log_assumptions"].extend([
        "总量 N=40（§11 默认值）",
        "负样本 6 条（§11 默认值）",
        "去重相似度阈值 0.85（同 section 桶内）",
        "G 列 embedding 语义补全默认关闭（§8.1 第3步可选项；关键词/子串匹配已足够）",
        "负样本不进行 embedding 去重",
        "G 列剔除 doc_summary 类（§5 规则4），除非该 chunk 是 primary_chunk_id 来源",
    ])

    log_event("[STEP2] LLM 生成可答题")
    answerable_rows = []
    for task in BUCKETS:
        if _stats["llm_calls"] >= MAX_LLM_CALLS:
            log_event("[STOP] LLM 调用上限触发，停止生成。")
            break
        log_event(f"  · 生成 {task['key']} {task['intent']}/{task['answer_type']} hint={task['hint'][:40]}")
        row = gen_one_qa(task, by_id, chunks)
        if row:
            answerable_rows.append(row)
    log_event(f"  共生成 {len(answerable_rows)} 条可答题（应得 {len(BUCKETS)}）")

    log_event("[STEP3] 构造负样本")
    negatives = gen_negatives()
    log_event(f"  共 {len(negatives)} 条负样本")

    rows = answerable_rows + negatives

    log_event("[STEP4] embedding 语义去重")
    n_before = len(rows)
    rows = dedup_by_embedding(rows)
    n_dedup = n_before - len(rows)
    drop_stats = {"dedup_dropped": n_dedup}

    log_event("[STEP5] 校验 §9")
    issues, short_ratio = validate_rows(rows, by_id, chunks)
    if issues:
        log_event(f"[VALIDATE] {len(issues)} 项问题：")
        for idx, msg in issues[:30]:
            log_event(f"  - row#{idx}: {msg}")
    log_event(f"[VALIDATE] question < 30 字比例：{short_ratio:.0%}")

    # 自动修复：丢弃严重不合规条目
    fixed_rows = []
    for r in rows:
        bad = False
        if r["answer_type"] != "unanswerable":
            gids = [g for g in r["gold_chunk_ids"].split(";") if g]
            f_ok = any(_norm(r["supporting_text"]) in _norm(by_id[g].get("content") or "") for g in gids if g in by_id)
            if not f_ok:
                bad = True
                log_event(f"[FIX] 丢弃 F 子串校验失败的条目：{r['question'][:30]}…")
            elif len(r["question"]) > 50:
                bad = True
                log_event(f"[FIX] 丢弃 question 过长条目：{r['question'][:30]}…")
        if not bad:
            fixed_rows.append(r)
    rows = fixed_rows
    log_event(f"  最终 {len(rows)} 条")

    log_event(f"[STEP6] 写出 xlsx → {OUT_XLSX}")
    write_xlsx(rows, OUT_XLSX)

    log_event(f"[STEP7] 写出运行日志 → {OUT_LOG}")
    write_log(rows, OUT_LOG, drop_stats)

    # 简短控制台总结
    print("\n=== DONE ===")
    print(f"rows = {len(rows)} → {OUT_XLSX}")
    print(f"log → {OUT_LOG}")


if __name__ == "__main__":
    main()
